#!python3

import openpyxl
import re

#------------------ここから上はいじらない-------------------------------



#listTextには誘導枠テキストを記入してください、「 "", 」の記載を忘れずに。
# 例；"「血圧が高めの方へ」トマト由来のGABAで出来る、飲む血圧対策",

#listImageには使用する誘導枠画像のURLを記載してください、
#URLは管理画面> by contentの時に画像の上にマウスを置いた時に出てくる画像のURLを使ってください

listText = [
    "「飲む血圧対策」トマト由来のGABAで美味しく血圧サポート",
    "高めの血圧に、______が作った「トマトドリンク」が凄い",
    "「血圧が高めの方へ」______のGABAトマトで美味しく対策",
    "続けたくなる血圧対策！あの______が作った「トマトドリンク」が話題",
    "______が開発「飲む血圧対策」美味しくて続けやすいと話題",
    "「高めの血圧に、2021年の新習慣」______のトマトドリンクが凄い",
    "「血圧130-139mmHgの方に」______のトマトドリンクが凄い"
]

listImage = [
    "http://images.outbrainimg.com/transform/v3/ey01.jpg",
    "http://images.outbrainimg.com/transform/v3/ey02.jpg",
    "http://images.outbrainimg.com/transform/v3/ey03.jpg",
    "http://images.outbrainimg.com/transform/v3/ey04.jpg",
]

#

#参照元のエクセルのファイル名（アドエビスからダウンロードしたもの）を入れてください
fromExel = "SPzerothree.xlsx"

#出力先のエクセルのファイル名を入れてください(任意で大丈夫ですが、最後に ".xlsx" を必ずつけてください)
toExel = "pc_one.xlsx"



#---------------------ここから下はいじらない-------------------------------------

wb = openpyxl.load_workbook(fromExel)
wb2paste = openpyxl.Workbook()

sheet = wb.active 
sheet2paste = wb2paste['Sheet']

sheet2paste.cell(row = 1, column = 1).value = 'URL(Required)'
sheet2paste.cell(row = 1, column = 2).value = 'Title'
sheet2paste.cell(row = 1, column = 3).value = 'ImageUrl'

numText = len(listText)
numImage = len(listImage)

totalNumCR = numText * numImage

#1 printing URL

for i in range(2,totalNumCR+2):
    var1 = sheet.cell(row=i, column=17).value

    urlFront = var1[0:var1.find('?')]
    urlBack = re.search(r'(?<=\?)[^.\s]*',var1).group(0)

    var2paste = urlFront +'?affiliate=0004011gbt&' + urlBack

    sheet2paste.cell(row = i, column = 1).value = var2paste

#2 printing text

var4row = 2
for b in range(numImage):
    for a in listText:
        sheet2paste.cell(row = var4row, column = 2).value = a
        var4row += 1

#3 printing imageURL

var4row = 2
for d in listImage:
    for c in range(numText):        
        sheet2paste.cell(row = var4row, column = 3).value = d
        var4row += 1
    
wb2paste.save(toExel)

print('Done.')